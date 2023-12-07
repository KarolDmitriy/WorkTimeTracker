import openpyxl

def read_and_convert_data(input_file, output_file):
    try:
        # Открываем файл daily_entries
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        # Создаем новый файл
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        # Добавляем заголовки
        headers = ["Табельный", "ФИО", "Дата входа", "Время входа", "Дата выхода", "Время выхода", "График"]
        new_sheet.append(headers)

        # Считываем и конвертируем строки
        for row in sheet.iter_rows(min_row=2, values_only=True):
            converted_row = [
                row[0],  # Табельный
                row[1],  # ФИО
                row[2],  # Дата входа (предполагается, что она уже в нужном формате)
                row[3],  # Время входа (предполагается, что оно уже в нужном формате)
                row[4],  # Дата выхода (предполагается, что она уже в нужном формате)
                row[5],  # Время выхода (предполагается, что оно уже в нужном формате)
                row[6],  # График
            ]

            new_sheet.append(converted_row)

        # Сохраняем новый файл
        new_workbook.save(output_file)

    except Exception as e:
        print(f"Ошибка при выполнении операций: {e}")

# Пример использования
input_file = "data/daily_entries.xlsx"
output_file = "data/new_format_daily_entries.xlsx"

read_and_convert_data(input_file, output_file)
