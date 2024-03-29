import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog

def check_plan_actual_time(comments_file, work_schedule_file, absence_file, start_date, end_date):
    # Загрузка данных из файла comments.xlsx
    comments_df = pd.read_excel(comments_file)

    # Открытие файла с табелем и чтение значений ячеек M1 и P1
    workbook = load_workbook(work_schedule_file)
    sheet = workbook.active
    month_name = sheet['M1'].value
    year = sheet['P1'].value

    # Загрузка данных из файла отсутствия
    absence_df = pd.read_excel(absence_file)

    # Словарь для соответствия месяцев на русском языке и их числового представления
    month_dict = {'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4, 'Май': 5, 'Июнь': 6,
                  'Июль': 7, 'Август': 8, 'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12}

    # Расчет количества дней в текущем месяце
    last_day_of_month = (datetime(year, month_dict[month_name], 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # Получение стартовой ячейки B13
    start_cell = sheet['B13'].value

    # Преобразование дат в столбце 'Дата входа' в формат datetime
    comments_df['Дата входа'] = pd.to_datetime(comments_df['Дата входа'], format='%d.%m.%Y', errors='coerce')

    # Итерация по табельным номерам (каждый 4-й ряд начиная с B13)
    for row_offset in range(sheet.min_row - 13, sheet.max_row - 12, 4):
        current_start_cell = sheet.cell(row=13 + row_offset, column=2).value

        # Проверяем наличие табельного номера в data frame
        if current_start_cell not in comments_df['Табельный'].values:
            # Проверяем, есть ли запись об отсутствии для данного табельного номера
            absence_record = absence_df[(absence_df['Таб.№'] == current_start_cell)]
            if not absence_record.empty:
                # Получаем букву из колонки "Вид отсутствия"
                absence_letter = absence_record.iloc[0]['Вид отсутствия']
                # Итерация по дням текущего месяца
                for day in range(1, last_day_of_month.day + 1):
                    current_date = datetime(year, month_dict[month_name], day)
                    # Проверяем, попадает ли текущая дата в период отсутствия
                    if (absence_record['Начало'].iloc[0] <= current_date <= absence_record['Истечение'].iloc[0]):
                        # Записываем букву в ячейку соответствующего дня в файле work_schedule_file,
                        # только если дата входит в заданный диапазон
                        if start_date <= current_date <= end_date:
                            sheet.cell(row=13 + row_offset, column=day + 5).value = absence_letter
                            # Закрашиваем ячейку
                            sheet.cell(row=13 + row_offset, column=day + 5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            continue  # Пропускаем выполнение, если табельный номер отсутствует
        else:
            # Ищем строки в data frame с таким же значением, как в стартовой ячейке
            matching_rows = comments_df[comments_df['Табельный'] == current_start_cell]

            # Итерация по дням текущего месяца
            for day in range(1, last_day_of_month.day + 1):
                current_date = datetime(year, month_dict[month_name], day)

                # Проверяем значение в ячейке F13
                f13_value = sheet.cell(row=13 + row_offset, column=day + 5).value

                # Проверяем, есть ли такой день в data frame
                matching_rows_for_date = matching_rows[matching_rows['Дата входа'].dt.date == current_date.date()]

                if not matching_rows_for_date.empty:
                    if pd.notna(f13_value):
                        matching_row = matching_rows_for_date

                        if not matching_row.empty and 'Вход вовремя - Выход вовремя' in matching_row.iloc[0]['Комментарий']:
                            continue
                        else:
                            # Неверный комментарий, закрасить ячейку F13 красным,
                            # только если дата входит в заданный диапазон
                            if start_date <= current_date <= end_date:
                                sheet.cell(row=13 + row_offset, column=day + 5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                # Проверяем, есть ли запись об отсутствии для данной даты и табельного номера
                                absence_record = absence_df[(absence_df['Таб.№'] == current_start_cell) & (absence_df['Начало'] <= current_date) & (absence_df['Истечение'] >= current_date)]
                                if not absence_record.empty:
                                    # Получаем букву из колонки "Вид отсутствия"
                                    absence_letter = absence_record.iloc[0]['Вид отсутствия']
                                    # Записываем букву в ячейку соответствующего дня в файле work_schedule_file,
                                    # только если дата входит в заданный диапазон
                                    sheet.cell(row=13 + row_offset, column=day + 5).value = absence_letter
                    else:
                        # Значение в F13 отсутствует, закрасить ячейку F13 красным,
                        # только если дата входит в заданный диапазон
                        if start_date <= current_date <= end_date:
                            sheet.cell(row=13 + row_offset, column=day + 5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            # Проверяем, есть ли запись об отсутствии для данной даты и табельного номера
                            absence_record = absence_df[(absence_df['Таб.№'] == current_start_cell) & (absence_df['Начало'] <= current_date) & (absence_df['Истечение'] >= current_date)]
                            if not absence_record.empty:
                                # Получаем букву из колонки "Вид отсутствия"
                                absence_letter = absence_record.iloc[0]['Вид отсутствия']
                                # Записываем букву в ячейку соответствующего дня в файле work_schedule_file,
                                # только если дата входит в заданный диапазон
                                sheet.cell(row=13 + row_offset, column=day + 5).value = absence_letter
                else:
                    if pd.notna(f13_value):
                        # День отсутствует в data frame, закрасить ячейку F13 красным,
                        # только если дата входит в заданный диапазон
                        if start_date <= current_date <= end_date:
                            sheet.cell(row=13 + row_offset, column=day + 5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            # Проверяем, есть ли запись об отсутствии для данной даты и табельного номера
                            absence_record = absence_df[(absence_df['Таб.№'] == current_start_cell) & (absence_df['Начало'] <= current_date) & (absence_df['Истечение'] >= current_date)]
                            if not absence_record.empty:
                                # Получаем букву из колонки "Вид отсутствия"
                                absence_letter = absence_record.iloc[0]['Вид отсутствия']
                                # Записываем букву в ячейку соответствующего дня в файле work_schedule_file,
                                # только если дата входит в заданный диапазон
                                sheet.cell(row=13 + row_offset, column=day + 5).value = absence_letter

    # Сохранение изменений в файле
    workbook.save(work_schedule_file)


def check_plan_actual_time_gui():
    root = tk.Tk()
    root.title("Check Plan and Actual Time")

    def browse_comments_file():
        comments_file_path = filedialog.askopenfilename(title="Выберите файл с комментариями", filetypes=[("Excel files", "*.xlsx")])
        comments_file_var.set(comments_file_path)

    def browse_work_schedule_file():
        work_schedule_file_path = filedialog.askopenfilename(title="Выберите файл с табелем", filetypes=[("Excel files", "*.xlsx")])
        work_schedule_file_var.set(work_schedule_file_path)

    def browse_absence_file():
        absence_file_path = filedialog.askopenfilename(title="Выберите файл отсутствия", filetypes=[("Excel files", "*.xlsx")])
        absence_file_var.set(absence_file_path)

    def run_check():
        comments_file_path = comments_file_var.get()
        work_schedule_file_path = work_schedule_file_var.get()
        absence_file_path = absence_file_var.get()

        start_date_str = start_date_var.get()  # Получаем начальную дату из виджета Entry
        end_date_str = end_date_var.get()  # Получаем конечную дату из виджета Entry

        if not comments_file_path or not work_schedule_file_path or not absence_file_path:
            result_label.config(text="Select all files", fg="red")
            return

        # Проверяем, что введены начальная и конечная даты
        if not start_date_str or not end_date_str:
            result_label.config(text="Enter start and end dates", fg="red")
            return

        # Преобразовываем строки с датами в объекты datetime
        start_date = datetime.strptime(start_date_str, '%d.%m.%Y')
        end_date = datetime.strptime(end_date_str, '%d.%m.%Y')

        # Проверяем, что начальная дата меньше или равна конечной
        if start_date > end_date:
            result_label.config(text="Start date must be before or equal to end date", fg="red")
            return

        check_plan_actual_time(comments_file_path, work_schedule_file_path, absence_file_path, start_date, end_date)
        result_label.config(text="Проверка выполнена!", fg="green")

    # Переменные для хранения путей к файлам
    comments_file_var = tk.StringVar()
    work_schedule_file_var = tk.StringVar()
    absence_file_var = tk.StringVar()

    # Метка и кнопка для выбора файла комментариев
    tk.Label(root, text="Выберите файл с комментариями:").grid(row=0, column=0)
    tk.Entry(root, textvariable=comments_file_var, state="readonly", width=50).grid(row=0, column=1)
    tk.Button(root, text="Browse", command=browse_comments_file).grid(row=0, column=2)

    # Метка и кнопка для выбора файла work_schedule_file
    tk.Label(root, text="Выберите файл с табелем:").grid(row=1, column=0)
    tk.Entry(root, textvariable=work_schedule_file_var, state="readonly", width=50).grid(row=1, column=1)
    tk.Button(root, text="Browse", command=browse_work_schedule_file).grid(row=1, column=2)


    # Метка и кнопка для выбора файла отсутствия
    tk.Label(root, text="Выберите файл отсутствия:").grid(row=2, column=0)
    tk.Entry(root, textvariable=absence_file_var, state="readonly", width=50).grid(row=2, column=1)
    tk.Button(root, text="Browse", command=browse_absence_file).grid(row=2, column=2)

    # Кнопка для запуска проверки
    tk.Button(root, text="Проверить", command=run_check).grid(row=3, column=0, columnspan=3, pady=10)

    # Метка для отображения результата
    result_label = tk.Label(root, text="", fg="black")
    result_label.grid(row=4, column=0, columnspan=3)

    # Добавляем новые метки и виджеты Entry для ввода дат
    tk.Label(root, text="Введите начальную дату (дд.мм.гггг):").grid(row=4, column=0)
    start_date_var = tk.StringVar()
    tk.Entry(root, textvariable=start_date_var, width=15).grid(row=4, column=1)

    tk.Label(root, text="Введите конечную дату (дд.мм.гггг):").grid(row=5, column=0)
    end_date_var = tk.StringVar()
    tk.Entry(root, textvariable=end_date_var, width=15).grid(row=5, column=1)

    root.mainloop()

check_plan_actual_time_gui()