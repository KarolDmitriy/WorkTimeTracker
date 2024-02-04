from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

def update_employee_absences(absences_file=None, output_file=None):
    if absences_file is None:
        # Если absences_file не предоставлен, позволяет пользователю выбрать файл
        absences_file = filedialog.askopenfilename(title="Выберите файл с отсутствиями", filetypes=[("Excel files", "*.xlsx")])

    if output_file is None:
        # Если output_file не предоставлен, позволяет пользователю выбрать файл
        output_file = filedialog.asksaveasfilename(title="Выберите файл для сохранения", filetypes=[("Excel files", "*.xlsx")])

    try:
        # Открываем указанный файл excel
        workbook = load_workbook(output_file)

        # Записываем значение ячейки F7 в переменную numberOfDays
        sheet = workbook.active
        numberOfDays = sheet.cell(row=7, column=6).value

        # Если ячейка F7 содержит целое число, используем его
        if not isinstance(numberOfDays, int):
            numberOfDays = 0

        # Открываем файл employee_absences
        absences_workbook = load_workbook(absences_file)
        absences_sheet = absences_workbook.active

        startRow = 13
        rowToName = 13

        for _ in range((sheet.max_row - startRow) // 4 + 1):
            # Записываем значения ячеек M1, P1, AQ13 в переменную monthYearGraph
            monthYearGraph = f"{sheet.cell(row=1, column=13).value} {sheet.cell(row=1, column=16).value} {sheet.cell(row=startRow, column=43).value}"

            # Находим ячейку с таким же значением, как в monthYearGraph
            absences_value_cell = None
            for row in absences_sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=absences_sheet.max_row):
                for cell in row:
                    if cell.value == monthYearGraph:
                        absences_value_cell = cell
                        break
                if absences_value_cell:
                    break

            # Если ячейка найдена, копируем диапазон значений
            if absences_value_cell:
                source_range = absences_sheet.iter_cols(min_col=absences_value_cell.column + 1,
                                                        max_col=absences_value_cell.column + numberOfDays,
                                                        min_row=absences_value_cell.row, max_row=absences_value_cell.row)
                values_to_paste = [cell[0].value for cell in source_range]

                # Вставляем диапазон в указанный файл excel
                for i, value in enumerate(values_to_paste):
                    # Проверяем, есть ли буквы в конце значения и является ли значение строкой
                    if value and isinstance(value, str) and value[-1].isalpha():
                        # Убираем букву в конце, заменяем запятую на точку, преобразуем в float и вставляем
                        value_without_letter = value[:-1].replace(',', '.')
                        sheet.cell(row=rowToName, column=i + 6, value=float(value_without_letter))
                        # В следующую ячейку вставляем цифру 8
                        sheet.cell(row=rowToName + 1, column=i + 6, value=8)
                    else:
                        # Если нет букв в конце или не является строкой, просто вставляем значение
                        sheet.cell(row=rowToName, column=i + 6, value=value)

            # Итерация по работникам
            startRow += 4
            rowToName += 4

        # Сохраняем изменения
        workbook.save(output_file)
        absences_workbook.save(absences_file)

    except Exception as e:
        print(f"Ошибка: {e}")

# Создаем простой графический интерфейс
def create_gui():
    root = tk.Tk()
    root.title("Update Employee Absences")

    def browse_absences_file():
        nonlocal absences_file_var
        absences_file_path = filedialog.askopenfilename(title="Выберите файл с отсутствиями", filetypes=[("Excel files", "*.xlsx")])
        absences_file_var.set(absences_file_path)

    def browse_output_file():
        nonlocal output_file_var
        output_file_path = filedialog.asksaveasfilename(title="Выберите файл для сохранения", filetypes=[("Excel files", "*.xlsx")])
        output_file_var.set(output_file_path)

    def run_update():
        absences_file_path = absences_file_var.get()
        output_file_path = output_file_var.get()

        if not absences_file_path or not output_file_path:
            result_label.config(text="Выберите все файлы", fg="red")
            return

        update_employee_absences(absences_file_path, output_file_path)
        result_label.config(text="Обновление выполнено!", fg="green")

    # Переменные для хранения путей к файлам
    absences_file_var = tk.StringVar()
    output_file_var = tk.StringVar()

    # Метка и кнопка для выбора absences_file
    tk.Label(root, text="Выберите файл с отсутствиями:").grid(row=0, column=0)
    tk.Entry(root, textvariable=absences_file_var, state="readonly", width=50).grid(row=0, column=1)
    tk.Button(root, text="Browse", command=browse_absences_file).grid(row=0, column=2)

    # Метка и кнопка для выбора output_file
    tk.Label(root, text="Выберите файл для сохранения:").grid(row=1, column=0)
    tk.Entry(root, textvariable=output_file_var, state="readonly", width=50).grid(row=1, column=1)
    tk.Button(root, text="Browse", command=browse_output_file).grid(row=1, column=2)

    # Кнопка для запуска обновления
    tk.Button(root, text="Обновить", command=run_update).grid(row=2, column=0, columnspan=3, pady=10)

    # Метка для отображения результата
    result_label = tk.Label(root, text="", fg="black")
    result_label.grid(row=3, column=0, columnspan=3)

    root.mainloop()

# Запускаем графический интерфейс
create_gui()
