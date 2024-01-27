import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, time

def format_date(date):
    if isinstance(date, datetime):
        return date.strftime("%d.%m.%Y")
    return None

def format_time(time_value):
    if isinstance(time_value, time):
        return time_value.strftime("%H:%M:%S")
    return None

def create_work_schedule_file(output_file, rows):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        header_row = ["Дата входа", "Время входа", "Дата выхода", "Время выхода"]
        sheet.append(header_row)

        for row in rows:
            sheet.append(row)

        workbook.save(output_file)

    except Exception as e:
        print(f"Ошибка при создании файла: {e}")

def split_work_schedules(input_file="data/work_schedules.xlsx", output_folder="data/graph"):
    try:
        schedules_workbook = openpyxl.load_workbook(input_file)
        schedules_sheet = schedules_workbook.active

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        current_schedule = None
        current_schedule_rows = []

        for row in schedules_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] != current_schedule:
                if current_schedule_rows:
                    output_file = f"{output_folder}/{current_schedule}.xlsx"
                    create_work_schedule_file(output_file, current_schedule_rows)

                current_schedule = row[0]
                current_schedule_rows = []

            formatted_row = [
                format_date(row[1]),
                format_time(row[2]),
                format_date(row[3]),
                format_time(row[4]),
            ]

            current_schedule_rows.append(formatted_row)

        if current_schedule_rows:
            output_file = f"{output_folder}/{current_schedule}.xlsx"
            create_work_schedule_file(output_file, current_schedule_rows)

        messagebox.showinfo("Успешно", "Файлы успешно разделены!")

    except Exception as e:
        print(f"Ошибка при выполнении операций: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")

def main():
    root = tk.Tk()
    root.withdraw()

    # Создаем главное окно
    main_window = tk.Tk()
    main_window.title("Разделение графиков")

    # Функция для обработки нажатия кнопки
    def process_button_click():
        # Запускаем функцию разделения графиков
        split_work_schedules(input_file_var.get(), output_folder_var.get())

    # Создаем и настраиваем виджеты
    input_file_label = tk.Label(main_window, text="Выберите файл с графиками:")
    input_file_label.pack()

    input_file_var = tk.StringVar()
    input_file_entry = tk.Entry(main_window, textvariable=input_file_var, state="readonly", width=50)
    input_file_entry.pack()

    def choose_input_file():
        input_file = filedialog.askopenfilename(title="Выберите файл",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        input_file_var.set(input_file)

    input_file_button = tk.Button(main_window, text="Обзор", command=choose_input_file)
    input_file_button.pack()

    output_folder_label = tk.Label(main_window, text="Выберите папку для сохранения:")
    output_folder_label.pack()

    output_folder_var = tk.StringVar()
    output_folder_entry = tk.Entry(main_window, textvariable=output_folder_var, state="readonly", width=50)
    output_folder_entry.pack()

    def choose_output_folder():
        output_folder = filedialog.askdirectory(title="Выберите папку для сохранения")
        output_folder_var.set(output_folder)

    output_folder_button = tk.Button(main_window, text="Обзор", command=choose_output_folder)
    output_folder_button.pack()

    process_button = tk.Button(main_window, text="Выполнить", command=process_button_click)
    process_button.pack()

    # Запускаем главный цикл обработки событий
    main_window.mainloop()

if __name__ == "__main__":
    main()
