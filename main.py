import openpyxl
import pandas as pd
from tqdm import tqdm
import logging
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog, messagebox

# Настройка логирования
logging.basicConfig(filename='app.log', level=logging.ERROR)

# Массив для хранения комментариев
comments_array = []

# Цвета ячеек для разных типов комментариев
COLOR_CHECK_ENTRY = "FFA500"  # Оранжевый
COLOR_MISSING_GRAPH = "FF0000"  # Красный

def format_date_in_row(row_list):
    return [
        row_list[0],  # Табельный
        row_list[1],  # ФИО
        row_list[2].strftime("%d.%m.%Y"),  # Дата входа
        row_list[3].strftime("%H:%M:%S"),  # Время входа
        row_list[4].strftime("%d.%m.%Y"),  # Дата выхода
        row_list[5].strftime("%H:%M:%S"),  # Время выхода
        row_list[6],  # График
    ]

def write_comment_to_array(comment, formatted_row, row_list):
    global comments_array
    comments_array.append([comment, row_list[0], row_list[1],
                           formatted_row[2], formatted_row[3], formatted_row[4],
                           formatted_row[5], row_list[6]])

def write_comments_to_excel(comment_file_path):
    try:
        # Загрузка существующего файла
        xls = openpyxl.load_workbook(comment_file_path)

        # Выбор нужного листа
        sheet = xls['Sheet1']

        # Определение следующей свободной строки
        next_row = sheet.max_row + 1

        for comment_row in comments_array:
            # Обновление значения в row_list[3] для текущей строки
            comment_row[3] = comment_row[3] if isinstance(comment_row[3], str) else comment_row[3].strftime("%d.%m.%Y")

            # Запись данных в новую строку
            for col_num, value in enumerate(comment_row, 1):
                cell = sheet.cell(row=next_row, column=col_num, value=value)

                # Проверка, содержится ли фраза "Проверить вход" в комментарии
                if "Проверить вход" in str(value) or "Проверить выход" in str(value) or "Выход в выходной день" in str(value) \
                        or "выход в выходной день" in str(value) or "Файл графика не найден" in str(value) \
                        or "вход по временному пропуску" in str(value):
                    # Применение цвета к ячейке
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            next_row += 1

        # Сохранение изменений
        xls.save(comment_file_path)
        xls.close()

    except Exception as e:
        error_message = f"Ошибка при записи комментариев в файл: {e}"
        logging.error(error_message)
        print(error_message)

def process_daily_entries(file_path):
    print("Запущена функция process_daily_entries")
    try:
        # Загрузка данных из файла daily_entries
        daily_entries = pd.read_excel(file_path, sheet_name='Лист1')

        # Создание директории, если она отсутствует
        output_dir = "data/graph"
        os.makedirs(output_dir, exist_ok=True)

        # Путь к файлу с комментариями
        comment_file_path = os.path.join(output_dir, "comments.xlsx")

        # Создание нового excel файла с комментариями, если он не существует
        if not os.path.exists(comment_file_path):
            df = pd.DataFrame(
                columns=["Комментарий", "Табельный", "ФИО", "Дата входа", "Время входа", "Дата выхода", "Время выхода",
                         "График"])
            df.to_excel(comment_file_path, index=False, engine='openpyxl', sheet_name='Sheet1')

        # Определение количества строк в файле
        limit = len(daily_entries)

        # Итерация по строкам из файла daily_entries
        for index, row in tqdm(daily_entries.iterrows(), total=len(daily_entries), desc="Обработка строк"):
            # Преобразование строки в список
            row_list = row.tolist()

            # Преобразование даты в нужный формат
            formatted_row = format_date_in_row(row_list)

            # Путь к файлу с графиком
            graph_file_path = os.path.join(output_dir, f"{row_list[-1]}.xlsx")

            # Проверка наличия файла с графиком
            if os.path.exists(graph_file_path):
                # Загрузка данных из файла с графиком
                graph_data = pd.read_excel(graph_file_path, engine='openpyxl')

                # Преобразование даты в нужный формат для сравнения
                formatted_date = row_list[2].strftime("%d.%m.%Y")
                formatted_date_exit = row_list[4].strftime("%d.%m.%Y")

                # Поиск строки по значению второго индекса из строки daily_entries
                matching_row = graph_data.loc[graph_data.iloc[:, 0] == formatted_date]
                matching_row_exit = graph_data.loc[graph_data.iloc[:, 2] == formatted_date_exit]

                # индекс строки из graph_data
                if not matching_row.empty and not matching_row_exit.empty:
                    valid_graphs = ['График 98 бригада 1', 'График 98 бригада 2',
                                    'График 1 бригада 1', 'График 1 бригада 2', 'График 1 бригада 3',
                                    'График 1 бригада 4', 'График 2 бригада 1', 'График 2 бригада 2',
                                    'График 2 бригада 3', 'График 2 бригада 4', 'График 3 бригада 1',
                                    'График 3 бригада 2', 'График 3 бригада 3', 'График 3 бригада 4',
                                    'График 5 бригада 1', 'График 5 бригада 2', 'График 5 бригада 3',
                                    'График 97 бригада 1', '5 дней 8 часов']

                    if row_list[6] in valid_graphs:
                        matching_time = matching_row.iloc[0, 1]
                        matching_time_exit = matching_row_exit.iloc[0, 3]

                        # Проверка входа
                        if pd.notna(matching_time) and formatted_row[3] <= matching_time:
                            comment = f"Вход вовремя"
                        elif pd.isna(matching_time):
                            comment = "Выход в выходной день"
                        else:
                            comment = f"Проверить вход"

                        # Проверка выхода
                        if not matching_row_exit.empty and pd.notna(matching_time_exit) and formatted_row[5] >= matching_time_exit:
                            comment += " - Выход вовремя"
                        elif pd.isna(matching_time_exit):
                            pass
                        else:
                            comment += " - Проверить выход"

                        # Добавление проверки на название графика
                        if row_list[6] == "5 дней 8 часов":
                            comment += " - вход по временному пропуску"
                    else:
                        comment = f"График не содержит информации для {row_list[2]}"
                else:
                    comment = f"Файл графика не найден для {row_list[-1]}"

                # Написать комментарий в массив
                write_comment_to_array(comment, formatted_row, row_list)

            # Условие для остановки после обработки limit строк
            if index + 1 >= limit:
                print(f"Обработано {limit} строк. Завершаем выполнение.")
                break

        # Записать комментарии в Excel после обработки всех строк
        write_comments_to_excel(comment_file_path)

    except Exception as e:
        error_message = f"Произошла ошибка: {e}"
        logging.error(error_message)
        print(error_message)


def main():
    root = tk.Tk()
    root.withdraw()

    input_file = filedialog.askopenfilename(title="Выберите файл",
                                             filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])

    if input_file:
        process_daily_entries(input_file)
        tk.messagebox.showinfo("Успешно", "Операция успешно выполнена!")

if __name__ == "__main__":
    main()
